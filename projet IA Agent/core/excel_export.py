from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment
import os

def export_session_to_excel(candidates, filename="report.xlsx"):
    wb = Workbook()
    
    # --- Summary Sheet ---
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    
    headers = ["Rang", "Nom", "Email", "Score Final", "Recommandation"]
    ws_summary.append(headers)
    
    for row in ws_summary["A1:E1"]:
        row.font = Font(bold=True, color="FFFFFF")
        row.fill = PatternFill("solid", fgColor="000000")
    
    # Sort candidates conceptually
    sorted_cands = sorted(candidates, key=lambda x: x.final_score or 0, reverse=True)
    
    for idx, c in enumerate(sorted_cands, 1):
        row = [idx, c.name, c.email, c.final_score, c.recommendation]
        ws_summary.append(row)
        
        # Color code score
        cell_score = ws_summary.cell(row=idx+1, column=4)
        if c.final_score and c.final_score > 80:
            cell_score.fill = PatternFill("solid", fgColor="92D050") # Green
        elif c.final_score and c.final_score < 50:
            cell_score.fill = PatternFill("solid", fgColor="FF0000") # Red
        else:
            cell_score.fill = PatternFill("solid", fgColor="FFC000") # Yellow
            
    # --- Candidate Detail Sheets ---
    for c in sorted_cands:
        safe_name = "".join([char for char in c.name if char.isalpha() or char.isdigit() or char==' ']).strip()[:30]
        if not safe_name: safe_name = f"Cand_{c.id}"
        
        ws_cand = wb.create_sheet(title=safe_name)
        ws_cand.append(["Profil Candidat", c.name])
        ws_cand.append(["Email", c.email])
        ws_cand.append(["Téléphone", c.phone])
        ws_cand.append([])
        ws_cand.append(["Note du Recruteur (IA)", c.recommendation])
        ws_cand.append([])
        ws_cand.append(["Dimension", "Score"])
        
        for d in c.dimension_scores:
            ws_cand.append([d.dimension_name, d.score])
            
    # --- Chart Sheet ---
    ws_chart = wb.create_sheet(title="Graphiques")
    chart = BarChart()
    chart.title = "Distribution des scores finaux"
    chart.y_axis.title = 'Score'
    chart.x_axis.title = 'Candidat'
    
    data = Reference(ws_summary, min_col=4, min_row=1, max_row=len(candidates)+1)
    cats = Reference(ws_summary, min_col=2, min_row=2, max_row=len(candidates)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws_chart.add_chart(chart, "B2")
    
    wb.save(filename)
    return filename
